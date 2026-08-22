"""FastAPI app: the web front end for the verification engine.

    python -m webapp          (or: uvicorn webapp.main:app --reload)

Single checks run inline. Bulk uploads run as background jobs so the browser
never waits on a 10,000-row list.
"""

import asyncio
import io
import os
import sys
from typing import Optional

from fastapi import (BackgroundTasks, FastAPI, File, Form, HTTPException,
                     Request, UploadFile)
from fastapi.responses import (FileResponse, HTMLResponse, JSONResponse,
                               RedirectResponse)
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from prefilter.cache import Cache            # noqa: E402
from prefilter.engine import run as run_engine  # noqa: E402
from webapp import auth                       # noqa: E402
from webapp.jobs import DONE, FAILED, RUNNING, JobStore  # noqa: E402

# Enable SMTP probing only where the sending IP is clean (i.e. the server).
# Off by default so a laptop run never emits false rejections.
ENABLE_SMTP = os.environ.get("ENABLE_SMTP", "").lower() in ("1", "true", "yes")
SMTP_HELO = os.environ.get("SMTP_HELO", "localhost")
# Empty sender (<>) avoids sender-callout verification rejecting our probes.
SMTP_MAIL_FROM = os.environ.get("SMTP_MAIL_FROM", "")
# Serve results from the verdict cache? Off => every check is fully live.
USE_CACHE = os.environ.get("USE_CACHE", "1").lower() in ("1", "true", "yes")

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DATA_DIR = os.path.join(os.path.dirname(BASE_DIR), "webdata")
os.makedirs(DATA_DIR, exist_ok=True)

CACHE_DB = os.path.join(DATA_DIR, "verdicts.sqlite3")
JOBS_DB = os.path.join(DATA_DIR, "jobs.sqlite3")
RESULTS_DIR = os.path.join(DATA_DIR, "results")

store = JobStore(JOBS_DB, RESULTS_DIR)
app = FastAPI(title="Email Verifier", docs_url=None, redoc_url=None)

# Fully public (no login): the account pages, docs, static, and the API (which
# authenticates by key inside the endpoint).
_PUBLIC_PATHS = {"/", "/account", "/signup", "/account/login", "/account/logout",
                 "/docs-api", "/favicon.ico", "/login", "/api/demo",
                 "/pricing", "/verify-email", "/forgot", "/reset"}
_PUBLIC_PREFIXES = ("/static/", "/api/v1/")


@app.middleware("http")
async def require_login(request: Request, call_next):
    path = request.url.path
    if path in _PUBLIC_PATHS or any(path.startswith(p) for p in _PUBLIC_PREFIXES):
        return await call_next(request)

    from webapp.api import current_account
    user = current_account(request)

    # Admin area needs an admin account.
    if path.startswith("/admin"):
        if user is None or not user.get("is_admin"):
            return RedirectResponse("/account", status_code=302)
        return await call_next(request)

    # Everything else needs any logged-in, enabled account.
    if user is not None and not user.get("disabled"):
        return await call_next(request)
    if path.startswith("/api/"):
        return JSONResponse({"detail": "login required"}, status_code=401)
    return RedirectResponse("/account", status_code=302)


@app.get("/login")
async def login_redirect():
    # Old owner login is superseded by accounts.
    return RedirectResponse("/account", status_code=302)


LOGIN_HTML = """<!DOCTYPE html><html lang="en"><head><meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>Sign in — Email Verifier</title><link rel="stylesheet" href="/static/style.css">
</head><body><div class="login-wrap"><form class="login-card" method="post" action="/login">
<div class="login-logo">✓</div><h1>Email Verifier</h1>
<p class="sub">Enter the password to continue.</p><!--NOTE-->
<input type="password" name="password" placeholder="Password" autofocus required>
<button type="submit">Sign in</button></form></div></body></html>"""


class VerifyRequest(BaseModel):
    email: str
    use_cache: bool = True


class FindRequest(BaseModel):
    name: str
    domain: str


def _cache():
    # None disables caching -- the engine then does no lookups and no writes,
    # so every check is live.
    return Cache(CACHE_DB) if USE_CACHE else None


# -------------------------------------------------------------- single ----

@app.post("/api/verify")
async def verify_one(req: VerifyRequest, request: Request):
    email = (req.email or "").strip()
    if not email:
        raise HTTPException(400, "email is required")

    from webapp.api import current_account, users as _users
    account = current_account(request)
    if account:
        ok, _rem = _users.consume(account["id"], 1)
        if not ok:
            raise HTTPException(429, "daily quota reached")

    smtp_config = None
    if ENABLE_SMTP:
        from prefilter.smtp_check import ProbeConfig
        smtp_config = ProbeConfig(helo=SMTP_HELO, mail_from=SMTP_MAIL_FROM)

    cache = _cache()
    try:
        verdicts, _report = await run_engine(
            [email], cache if req.use_cache else None, use_microsoft=True,
            use_smtp=ENABLE_SMTP, smtp_config=smtp_config)
    finally:
        if cache: cache.close()

    v = verdicts[0]
    if account:
        _users.log_query(account["id"], "verify", email, v.status, "web")
    return {
        "email": v.email,
        "status": v.status,
        "confidence": v.confidence,
        "disposition": v.disposition,
        "tier": v.tier or "vendor",
        "route": v.route,
        "reason": v.reason,
        "suggestion": v.suggestion,
        "billable": v.disposition == "to_vendor",
    }


# ---------------------------------------------------------------- find ----

@app.post("/api/find")
async def find_email(req: FindRequest, request: Request):
    from prefilter import finder

    name = (req.name or "").strip()
    domain = (req.domain or "").strip()
    if not name or not domain:
        raise HTTPException(400, "name and domain are both required")

    from webapp.api import current_account, users as _users
    account = current_account(request)
    if account:
        ok, _rem = _users.consume(account["id"], 1)
        if not ok:
            raise HTTPException(429, "daily quota reached")

    smtp_config = None
    if ENABLE_SMTP:
        from prefilter.smtp_check import ProbeConfig
        smtp_config = ProbeConfig(helo=SMTP_HELO, mail_from=SMTP_MAIL_FROM)

    cache = _cache()
    try:
        r = await finder.find(name, domain, cache=cache, use_microsoft=True,
                              use_smtp=ENABLE_SMTP, smtp_config=smtp_config)
    finally:
        if cache: cache.close()
    if account:
        _users.log_query(account["id"], "find", "%s @ %s" % (name, domain),
                         r.status, "web")
    return {
        "email": r.email, "status": r.status, "confidence": r.confidence,
        "method": r.method, "tried": r.tried, "candidates": r.candidates,
    }


# ---------------------------------------------------------------- bulk ----

def _looks_like_email(cell: str) -> bool:
    cell = (cell or "").strip()
    return "@" in cell and " " not in cell and "." in cell.split("@")[-1]


def _parse_table(raw: bytes, filename: str):
    """Read the whole sheet, not just the emails.

    Returns (header, rows, email_idx):
      * header    -- the column names if the file has a header row, else None
      * rows      -- every data row as a list of strings (all columns kept)
      * email_idx -- which column holds the addresses

    Keeping the full table is what lets the download echo the user's original
    columns back and append the verdict columns after them.
    """
    import csv as _csv

    name = filename.lower()
    if name.endswith((".xlsx", ".xlsm")):
        from openpyxl import load_workbook
        workbook = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
        rows = [["" if c is None else str(c).strip() for c in row]
                for row in workbook.active.iter_rows(values_only=True)]
        workbook.close()
    else:
        text = raw.decode("utf-8-sig", "replace")
        rows = [[(c or "").strip() for c in row]
                for row in _csv.reader(io.StringIO(text))]

    # Drop fully-empty rows (trailing blanks are common in exports).
    rows = [r for r in rows if any(cell for cell in r)]
    if not rows:
        return None, [], 0

    width = max(len(r) for r in rows)
    rows = [r + [""] * (width - len(r)) for r in rows]  # pad ragged rows

    # The email column is the one with the most address-shaped cells across the
    # body of the file (skip row 0 in case it is a header).
    body = rows[1:] if len(rows) > 1 else rows
    best_idx, best_hits = 0, -1
    for idx in range(width):
        hits = sum(1 for r in body if _looks_like_email(r[idx]))
        if hits > best_hits:
            best_idx, best_hits = idx, hits

    # Header present if row 0's email cell is NOT an address but the body has
    # them -- i.e. row 0 is labels, not data.
    header = None
    if len(rows) > 1 and not _looks_like_email(rows[0][best_idx]) and best_hits > 0:
        header = rows[0]
        data = rows[1:]
    else:
        data = rows

    return header, data, best_idx


def _extract_emails(raw: bytes, filename: str):
    """Flat list of addresses, for callers that only need the emails."""
    _header, rows, idx = _parse_table(raw, filename)
    return [r[idx] for r in rows if idx < len(r) and _looks_like_email(r[idx])]


async def _run_job(job_id: str, emails, pattern_threshold: int = 0):
    store.update(job_id, status=RUNNING, rows_in=len(emails))
    cache = _cache()
    try:
        def on_progress(stage, done, total):
            store.update(job_id, stage=stage, done=done, total=total)

        smtp_config = None
        if ENABLE_SMTP:
            from prefilter.smtp_check import ProbeConfig
            smtp_config = ProbeConfig(helo=SMTP_HELO, mail_from=SMTP_MAIL_FROM)

        verdicts, report = await run_engine(
            emails, cache, use_microsoft=True, use_smtp=ENABLE_SMTP,
            smtp_config=smtp_config, use_patterns=True,
            pattern_threshold=pattern_threshold, log=lambda *_a: None,
            on_progress=on_progress)
        store.write_results(job_id, verdicts)

        import json as _json
        counts = {}
        for v in verdicts:
            counts[v.status] = counts.get(v.status, 0) + 1
        store.update(job_id, status=DONE, stage="Finished",
                     unique_in=report.unique, resolved=report.siphoned,
                     billable=report.billable, counts=_json.dumps(counts),
                     finished_at=__import__("datetime").datetime.now(
                         __import__("datetime").timezone.utc
                     ).isoformat(timespec="seconds"))
    except Exception as exc:  # noqa: BLE001
        store.update(job_id, status=FAILED, error=str(exc)[:400])
    finally:
        if cache: cache.close()


@app.post("/api/bulk")
async def bulk(background: BackgroundTasks, file: UploadFile = File(...),
               threshold: int = 0):
    raw = await file.read()
    if len(raw) > 25 * 1024 * 1024:
        raise HTTPException(413, "file too large (max 25 MB)")
    try:
        header, rows, email_idx = _parse_table(raw, file.filename or "list.csv")
        emails = [r[email_idx] for r in rows
                  if email_idx < len(r) and _looks_like_email(r[email_idx])]
    except Exception as exc:  # noqa: BLE001
        raise HTTPException(400, "could not read the file: %s" % exc)
    if not emails:
        raise HTTPException(400, "no email addresses found in that file")

    threshold = max(0, min(100, int(threshold)))
    job_id = store.create(file.filename or "list.csv")
    # Keep the user's original columns so the download can echo them back.
    store.write_source(job_id, header, rows, email_idx)
    store.update(job_id, rows_in=len(emails))
    # Pass the coroutine FUNCTION, not a coroutine object -- FastAPI inspects
    # it and awaits async callables on the loop. Handing it a sync callable
    # would run it in a worker thread with no event loop.
    background.add_task(_run_job, job_id, emails, threshold)
    return {"job_id": job_id, "found": len(emails)}


def _extract_name_domain(raw: bytes, filename: str):
    """Pull (name, domain) pairs from a CSV/XLSX. Detects the two columns by
    header; falls back to first two columns."""
    import csv as _csv

    name = filename.lower()
    if name.endswith((".xlsx", ".xlsm")):
        from openpyxl import load_workbook
        workbook = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
        rows = [["" if c is None else str(c).strip() for c in row]
                for row in workbook.active.iter_rows(values_only=True)]
        workbook.close()
    else:
        rows = list(_csv.reader(io.StringIO(raw.decode("utf-8-sig", "replace"))))
    rows = [r for r in rows if any((c or "").strip() for c in r)]
    if not rows:
        return []

    header = [c.lower().strip() for c in rows[0]]
    name_i = domain_i = None
    for i, h in enumerate(header):
        if name_i is None and "name" in h:
            name_i = i
        if domain_i is None and ("domain" in h or "company" in h or "website" in h):
            domain_i = i
    body = rows[1:]
    if name_i is None or domain_i is None:
        name_i, domain_i, body = 0, 1, rows  # headerless: first two columns

    pairs = []
    for r in body:
        if len(r) <= max(name_i, domain_i):
            continue
        nm = (r[name_i] or "").strip()
        dom = (r[domain_i] or "").strip().lower().lstrip("@")
        dom = dom.replace("http://", "").replace("https://", "").split("/")[0]
        if nm and "." in dom:
            pairs.append((nm, dom))
    return pairs


async def _run_find_job(job_id: str, pairs):
    store.update(job_id, status=RUNNING, rows_in=len(pairs))
    cache = _cache()
    try:
        from prefilter import finder

        smtp_config = None
        if ENABLE_SMTP:
            from prefilter.smtp_check import ProbeConfig
            smtp_config = ProbeConfig(helo=SMTP_HELO, mail_from=SMTP_MAIL_FROM)

        def on_progress(done, total):
            store.update(job_id, stage="Finding emails", done=done, total=total)

        results = await finder.find_many(
            pairs, cache=cache, use_microsoft=True, use_smtp=ENABLE_SMTP,
            smtp_config=smtp_config, progress=on_progress)
        store.write_find_results(job_id, results)

        import json as _json
        counts = {}
        for r in results:
            counts[r.status] = counts.get(r.status, 0) + 1
        found = sum(1 for r in results if r.status == "found")
        store.update(job_id, status=DONE, stage="Finished",
                     unique_in=len(results), resolved=found,
                     billable=len(results) - found, counts=_json.dumps(counts),
                     finished_at=__import__("datetime").datetime.now(
                         __import__("datetime").timezone.utc
                     ).isoformat(timespec="seconds"))
    except Exception as exc:  # noqa: BLE001
        store.update(job_id, status=FAILED, error=str(exc)[:400])
    finally:
        if cache: cache.close()


@app.post("/api/bulk-find")
async def bulk_find(background: BackgroundTasks, file: UploadFile = File(...)):
    raw = await file.read()
    if len(raw) > 25 * 1024 * 1024:
        raise HTTPException(413, "file too large (max 25 MB)")
    try:
        pairs = _extract_name_domain(raw, file.filename or "list.csv")
    except Exception as exc:  # noqa: BLE001
        raise HTTPException(400, "could not read the file: %s" % exc)
    if not pairs:
        raise HTTPException(
            400, "no name + domain pairs found -- need a name column and a "
                 "domain column")

    job_id = store.create(file.filename or "find.csv")
    store.update(job_id, rows_in=len(pairs))
    background.add_task(_run_find_job, job_id, pairs)
    return {"job_id": job_id, "found": len(pairs)}


@app.get("/api/jobs")
async def list_jobs():
    return {"jobs": store.list()}


@app.get("/api/history")
async def my_history(request: Request):
    """This account's single verify + find checks, newest first.

    Bulk uploads live in /api/jobs; this is everything the History tab was
    missing -- the one-off checks the user ran from the Single and Find tools.
    """
    from webapp.api import current_account, users as _users
    account = current_account(request)
    if account is None:
        return {"checks": []}
    rows = _users.recent_queries(200, user_id=account["id"])
    checks = [{"kind": r["kind"], "query": r["query"], "result": r["result"],
               "via": r["via"], "at": r["at"]}
              for r in rows if r["kind"] in ("verify", "find")]
    return {"checks": checks}


@app.get("/api/jobs/{job_id}")
async def get_job(job_id: str):
    job = store.get(job_id)
    if job is None:
        raise HTTPException(404, "no such job")
    return job


@app.delete("/api/jobs/{job_id}")
async def delete_job(job_id: str):
    if not store.delete(job_id):
        raise HTTPException(404, "no such job")
    return {"deleted": job_id}


@app.get("/api/jobs/{job_id}/results")
async def job_results(job_id: str, which: str = "all", status: str = "",
                      limit: int = 500, offset: int = 0):
    if store.get(job_id) is None:
        raise HTTPException(404, "no such job")
    return store.read_results(job_id, which, status, limit, offset)


@app.get("/api/jobs/{job_id}/download")
async def download(job_id: str, which: str = "clearout"):
    job = store.get(job_id)
    if job is None:
        raise HTTPException(404, "no such job")
    # Prefer the enriched file (original columns + verdicts); fall back to the
    # plain verdict file for jobs created before this existed.
    path = store.enriched_path(job_id, which)
    if not os.path.exists(path):
        path = store.result_path(job_id, which)
    if not os.path.exists(path):
        raise HTTPException(404, "results not ready")
    stem = os.path.splitext(job["filename"])[0][:40]
    return FileResponse(path, media_type="text/csv",
                        filename="%s_%s.csv" % (stem, which))


@app.get("/api/me")
async def me(request: Request):
    """Current account summary — powers the sidebar quota widget."""
    from webapp.api import current_account
    from webapp.users import _today
    account = current_account(request)
    if account is None:
        raise HTTPException(401, "login required")
    used = account["used_today"] if account["quota_date"] == _today() else 0
    return {"email": account["email"], "plan": account["plan"],
            "used_today": used, "daily_quota": account["daily_quota"],
            "is_admin": bool(account["is_admin"]), "verified": bool(account["verified"])}


@app.get("/api/stats")
async def stats():
    cache = _cache()
    if cache is None:
        return {"cached_verdicts": 0, "cached_domains": 0, "breakdown": {},
                "caching": "off (live mode)"}
    try:
        verdicts, domains = cache.counts()
        return {"cached_verdicts": verdicts, "cached_domains": domains,
                "breakdown": cache.status_breakdown()}
    finally:
        cache.close()


# --------------------------------------------------------- public API ----

from webapp import api as public_api  # noqa: E402


def _smtp_config_factory():
    if not ENABLE_SMTP:
        return None
    from prefilter.smtp_check import ProbeConfig
    return ProbeConfig(helo=SMTP_HELO, mail_from=SMTP_MAIL_FROM)


def _finder_module():
    from prefilter import finder
    return finder


public_api.configure(run_engine=run_engine, finder=__import__(
    "prefilter.finder", fromlist=["find"]), cache_factory=_cache,
    smtp_config_factory=_smtp_config_factory, enable_smtp=ENABLE_SMTP)
app.include_router(public_api.router)

# Seed the owner's admin account from env (ADMIN_EMAIL + APP_PASSWORD).
_ADMIN_EMAIL = os.environ.get("ADMIN_EMAIL", "admin@xomexo.com")
if os.environ.get("APP_PASSWORD"):
    public_api.users.seed_admin(_ADMIN_EMAIL, os.environ["APP_PASSWORD"])


@app.get("/docs-api", response_class=HTMLResponse)
async def docs_api():
    from webapp import shell

    def block(code):
        return '<div class="doc-block"><pre>%s</pre></div>' % code

    body = ('<div class="pub-head"><div class="eyebrow">Developers</div>'
            '<h1>API documentation</h1>'
            '<p class="lead">Send your key as an <code>X-API-Key</code> header on every '
            'request. Free plan: 100 checks a day.</p></div>'
            '<h3>Verify one address</h3>' + block(
                "POST /api/v1/verify\n"
                "X-API-Key: your_key\n"
                "Content-Type: application/json\n\n"
                '{"email": "someone@company.com"}\n\n'
                '&rarr; {"email":"...", "status":"valid", "checked_by":"microsoft",\n'
                '     "quota_remaining":99}') +
            '<h3>Find an address</h3>' + block(
                "POST /api/v1/find\nX-API-Key: your_key\n\n"
                '{"name": "John Smith", "domain": "company.com"}\n\n'
                '&rarr; {"email":"john.smith@company.com", "status":"found"}') +
            '<h3>Verify a list</h3>' + block(
                "POST /api/v1/bulk\nX-API-Key: your_key\n\n"
                '{"emails": ["a@x.com", "b@y.com"]}      // up to 1000 per call\n\n'
                '&rarr; {"count":2, "resolved":1, "results":[ ... ]}') +
            '<h3>Check your usage</h3>' + block(
                "GET /api/v1/usage\nX-API-Key: your_key\n\n"
                '&rarr; {"plan":"free","used_today":3,"daily_quota":100,"remaining":97}') +
            '<h3>Statuses</h3><div class="legend-grid" style="max-width:560px">'
            '<div><span class="pill valid">valid</span> Confirmed real.</div>'
            '<div><span class="pill invalid">invalid</span> Confirmed undeliverable.</div>'
            '<div><span class="pill catch_all">catch all</span> Domain accepts everything.</div>'
            '<div><span class="pill unknown">unknown</span> Could not be proven.</div></div>'
            '<p class="muted" style="margin-top:22px">Errors: <code>401</code> bad key · '
            '<code>429</code> quota exhausted · <code>413</code> batch too large.</p>')
    return HTMLResponse(shell.public_page("API docs", body, active="docs"))


# ----------------------------------------------------------------- ui ----

STATIC_DIR = os.path.join(BASE_DIR, "static")
app.mount("/static", StaticFiles(directory=STATIC_DIR), name="static")


@app.get("/", response_class=HTMLResponse)
async def landing(request: Request):
    # Logged-in visitors skip the marketing page and go straight to the app.
    from webapp.api import current_account
    if current_account(request) is not None:
        return RedirectResponse("/app", status_code=302)
    with open(os.path.join(STATIC_DIR, "landing.html"), encoding="utf-8") as handle:
        return HTMLResponse(handle.read())


@app.post("/api/demo")
async def demo(req: VerifyRequest):
    """Anonymous try-it on the landing page. HTTPS tiers only (no SMTP), so a
    stranger can never touch the server IP, and globally rate-limited."""
    email = (req.email or "").strip()
    if "@" not in email:
        raise HTTPException(400, "enter a valid email")
    from webapp.api import users as _users
    if not _users.bump_counter("demo", 300, 3600):
        raise HTTPException(429, "demo busy — create a free account to keep going")
    cache = _cache()
    try:
        verdicts, _ = await run_engine([email], cache, use_microsoft=True,
                                       use_smtp=False)
    finally:
        if cache:
            cache.close()
    v = verdicts[0]
    return {"email": v.email, "status": v.status, "reason": v.reason}


@app.get("/app", response_class=HTMLResponse)
async def app_page(request: Request):
    from webapp.api import current_account
    account = current_account(request)
    if account is None:
        return RedirectResponse("/account", status_code=302)
    from webapp import shell
    from webapp.users import _today
    account["_used"] = (account["used_today"]
                        if account["quota_date"] == _today() else 0)
    with open(os.path.join(STATIC_DIR, "index.html"), encoding="utf-8") as handle:
        html = handle.read()
    # One sidebar, rendered from the same place as every other page.
    return HTMLResponse(html.replace("<!--SIDEBAR-->",
                                     shell.sidebar(account, active="single")))
