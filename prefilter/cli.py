"""Command line interface.

    prefilter run    raw_list.csv          -- shrink a list before uploading
    prefilter run    raw_list.csv --shadow -- upload everything, grade ourselves
    prefilter ingest clearout_results.csv  -- bank the paid verdicts
    prefilter stats                        -- cache size and shadow accuracy
"""

import argparse
import csv
import glob
import os
import sys
from collections import Counter
from typing import Dict, List, Optional, Tuple

from .cache import Cache
from .normalize import parse
from .pipeline import ESCALATE, RESOLVED, REVIEW, run_pipeline

DEFAULT_DB = "prefilter.sqlite3"

# Clearout's vocabulary, normalised to ours.
VENDOR_STATUS_MAP = {
    "valid": "valid", "deliverable": "valid", "ok": "valid",
    "invalid": "invalid", "undeliverable": "invalid",
    "catch_all": "catch_all", "catchall": "catch_all", "catch all": "catch_all",
    "accept_all": "catch_all", "acceptall": "catch_all", "accept all": "catch_all",
    "unknown": "unknown", "risky": "unknown", "abuse": "unknown",
    "disposable": "disposable", "role": "role",
}


def _parse_status_map(pairs: Optional[List[str]]) -> Dict[str, str]:
    """Turn --status-map "Not Validated=invalid" arguments into a lookup."""
    mapping: Dict[str, str] = {}
    for pair in pairs or []:
        if "=" not in pair:
            raise SystemExit(
                "--status-map expects LABEL=status, got {0!r}".format(pair))
        label, status = pair.split("=", 1)
        mapping[label.strip().lower()] = status.strip().lower()
    return mapping


def _resolve_status(raw_value: str, custom: Dict[str, str],
                    empty_status: Optional[str]) -> str:
    """Map an export's own vocabulary onto ours.

    A blank cell only means 'valid' if the exporter guarantees every row was
    processed. Absent that guarantee it stays unknown, because a blank is
    indistinguishable from a row that silently failed -- and trusting it would
    put an unverified address in front of a client's sending domain.
    """
    value = (raw_value or "").strip()
    if not value:
        return empty_status or "unknown"
    key = value.lower()
    if key in custom:
        return custom[key]
    return VENDOR_STATUS_MAP.get(key, key.replace(" ", "_"))


def _detect_column(header: List[str], candidates: List[str]) -> Optional[int]:
    lowered = [(cell or "").strip().lower() for cell in header]
    for index, cell in enumerate(lowered):
        if cell in candidates:
            return index
    for index, cell in enumerate(lowered):
        if any(candidate in cell for candidate in candidates):
            return index
    return None


def _read_rows(path: str) -> List[List[str]]:
    """Read a CSV or XLSX into plain rows of strings."""
    if path.lower().endswith((".xlsx", ".xlsm")):
        try:
            from openpyxl import load_workbook
        except ImportError:
            raise SystemExit(
                "reading .xlsx needs openpyxl -- pip install openpyxl")
        workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            rows = [
                ["" if cell is None else str(cell).strip() for cell in row]
                for row in workbook.active.iter_rows(values_only=True)
            ]
        finally:
            workbook.close()
        # Spreadsheets routinely carry trailing blank rows.
        while rows and not any(cell for cell in rows[-1]):
            rows.pop()
        return rows
    with open(path, "r", encoding="utf-8-sig", errors="replace", newline="") as handle:
        return list(csv.reader(handle))


def _read_column(path: str, candidates: List[str],
                 explicit: Optional[str] = None) -> Tuple[List[str], List[List[str]], int]:
    rows = _read_rows(path)
    if not rows:
        raise SystemExit("{0} is empty".format(path))

    header, body = rows[0], rows[1:]
    if explicit is not None:
        if explicit.isdigit():
            index = int(explicit)
        else:
            index = _detect_column(header, [explicit.lower()])
            if index is None:
                raise SystemExit("no column named {0!r} in {1}".format(explicit, path))
    else:
        index = _detect_column(header, candidates)

    if index is None:
        if len(header) == 1:
            index = 0
        else:
            raise SystemExit(
                "could not find a column matching {0} in {1} -- pass --column"
                .format("/".join(candidates), path))

    # A headerless single-column file would otherwise lose its first address.
    if "@" in (header[index] or ""):
        body = rows
    return header, body, index


def cmd_run(args: argparse.Namespace) -> int:
    header, body, index = _read_column(
        args.input, ["email", "email address", "e-mail", "mail"], args.column)
    emails = [row[index] for row in body if len(row) > index and (row[index] or "").strip()]

    disposable = None
    if args.disposable_file:
        with open(args.disposable_file, "r", encoding="utf-8", errors="replace") as handle:
            disposable = {line.strip().lower() for line in handle if line.strip()}

    cache = Cache(args.db)
    rows, report = run_pipeline(
        emails, cache,
        ttl_days=args.ttl, domain_ttl_days=args.domain_ttl,
        concurrency=args.concurrency, timeout=args.timeout,
        drop_role=args.drop_role, disposable=disposable,
        nameservers=args.nameserver or None)

    os.makedirs(args.outdir, exist_ok=True)
    upload_path = os.path.join(args.outdir, "upload_to_clearout.csv")
    resolved_path = os.path.join(args.outdir, "resolved_locally.csv")
    review_path = os.path.join(args.outdir, "review.csv")

    fields = ["email", "canonical", "domain", "status", "reason", "source",
              "gateway", "provider", "is_role", "suggestion"]

    def write(path: str, subset: List) -> None:
        with open(path, "w", encoding="utf-8", newline="") as handle:
            writer = csv.writer(handle)
            writer.writerow(fields)
            for row in subset:
                writer.writerow([
                    row.original, row.canonical, row.domain, row.status,
                    row.reason, row.source, row.gateway, row.provider,
                    "yes" if row.is_role else "", row.suggestion])

    if args.shadow:
        # Shadow mode uploads everything we could parse, including the
        # addresses we would have dropped, so Clearout adjudicates our drops
        # and we get a real false-negative rate instead of an assumption.
        upload = [row for row in rows if row.canonical and row.status != "duplicate"]
    else:
        upload = [row for row in rows if row.bucket == ESCALATE]

    write(upload_path, upload)
    write(resolved_path, [row for row in rows if row.bucket == RESOLVED])
    write(review_path, [row for row in rows if row.bucket == REVIEW])

    billable_before = report.unique
    billable_after = len(upload)
    saved = max(0, billable_before - billable_after)

    print("input rows              : {0}".format(report.total_input))
    print("unique addresses        : {0}".format(report.unique))
    print("duplicates removed      : {0}".format(report.duplicates_removed))
    print("resolved locally        : {0}".format(report.counts[RESOLVED]))
    print("needs human review      : {0}".format(report.counts[REVIEW]))
    print("to upload to Clearout   : {0}{1}".format(
        billable_after, "  (SHADOW: full list)" if args.shadow else ""))
    if not args.shadow:
        print("credits avoided         : {0}  (~{1}{2:.2f})".format(
            saved, args.currency, saved * args.credit_price))
    print()
    print("top reasons:")
    for reason, count in report.reasons.most_common(12):
        print("  {0:>7}  {1}".format(count, reason))
    if report.gateways:
        print()
        print("security gateways seen (these must stay on the paid tier):")
        for gateway, count in report.gateways.most_common():
            print("  {0:>7}  {1}".format(count, gateway))
    print()
    print("wrote {0}".format(upload_path))
    print("      {0}".format(resolved_path))
    print("      {0}".format(review_path))

    cache.close()
    return 0


def cmd_ingest(args: argparse.Namespace) -> int:
    cache = Cache(args.db)
    status_map = _parse_status_map(args.status_map)
    seen_labels: Counter = Counter()
    # An address can appear twice in one export carrying different labels.
    # Silently letting the last row win would decide it by row order, so a
    # conflict downgrades to 'unknown' and gets re-verified instead.
    decided: Dict[str, str] = {}
    conflicted: set = set()
    paths: List[str] = []
    for pattern in args.input:
        matched = glob.glob(pattern)
        paths.extend(matched if matched else [pattern])

    total = 0
    compared = 0
    for path in paths:
        if not os.path.exists(path):
            print("skipping missing file: {0}".format(path), file=sys.stderr)
            continue
        header, body, email_index = _read_column(
            path, ["email", "email address", "e-mail", "mail"], args.column)
        if args.status_column is not None:
            status_index = (int(args.status_column)
                            if str(args.status_column).isdigit()
                            else _detect_column(header, [str(args.status_column).lower()]))
        else:
            status_index = _detect_column(header, ["status", "result", "verdict"])
        sub_index = _detect_column(header, ["sub status", "sub_status", "substatus", "reason"])
        if status_index is None:
            print("no status column in {0}, skipping -- pass --status-column".format(path),
                  file=sys.stderr)
            continue

        for raw_row in body:
            # Only the email cell is required. A trailing status cell that is
            # blank is often omitted from the row entirely by the exporter, so
            # a short row must not silently drop the address.
            if len(raw_row) <= email_index:
                continue
            raw_email = (raw_row[email_index] or "").strip()
            if not raw_email:
                continue
            parsed = parse(raw_email)
            if not parsed.canonical:
                continue

            vendor_cell = raw_row[status_index] if len(raw_row) > status_index else ""
            vendor_status = _resolve_status(vendor_cell, status_map, args.empty_status)
            seen_labels[(vendor_cell or "").strip() or "(blank)"] += 1

            prior = decided.get(parsed.canonical)
            if prior is not None and prior != vendor_status:
                vendor_status = "unknown"
                conflicted.add(parsed.canonical)
            decided[parsed.canonical] = vendor_status

            sub_status = ""
            if sub_index is not None and len(raw_row) > sub_index:
                sub_status = (raw_row[sub_index] or "").strip()

            # Read our own verdict BEFORE overwriting it with the vendor's.
            local = cache.conn.execute(
                "SELECT status, source FROM verdicts WHERE canonical = ?",
                (parsed.canonical,)).fetchone()
            if (local is not None and not args.dry_run
                    and str(local["source"]).startswith("prefilter")):
                domain_row = cache.conn.execute(
                    "SELECT gateway, provider FROM domains WHERE domain = ?",
                    (parsed.domain,)).fetchone()
                cache.log_shadow(
                    parsed.canonical, parsed.domain,
                    domain_row["gateway"] if domain_row else "",
                    domain_row["provider"] if domain_row else "",
                    local["status"], vendor_status)
                compared += 1

            if not args.dry_run:
                cache.put(parsed.canonical, raw_email, "clearout", vendor_status,
                          sub_status)
            total += 1
        if not args.dry_run:
            cache.commit()
        print("{0} {1}".format("inspected" if args.dry_run else "ingested", path))

    verdicts, domains = cache.counts()
    print()
    if seen_labels:
        print("labels found in the status column:")
        for label, count in seen_labels.most_common():
            print("  {0:>8}  {1!r} -> {2}".format(
                count, label,
                _resolve_status("" if label == "(blank)" else label,
                                status_map, args.empty_status)))
        print()
    print("rows processed          : {0}".format(total))
    print("unique addresses        : {0}".format(len(decided)))
    if conflicted:
        print("addresses w/ conflicting labels : {0}  -> forced to 'unknown', "
              "will be re-verified".format(len(conflicted)))
    print("shadow comparisons      : {0}".format(compared))
    print("cache now holds         : {0} verdicts, {1} domains".format(verdicts, domains))
    cache.close()
    return 0


def cmd_engine(args: argparse.Namespace) -> int:
    """Run the full tiered engine and split the list into siphoned vs billable."""
    import asyncio as _asyncio

    from .engine import REVIEW, SIPHONED, TO_VENDOR, run

    _header, body, index = _read_column(
        args.input, ["email", "email address", "e-mail", "mail"], args.column)
    emails = [row[index] for row in body
              if len(row) > index and (row[index] or "").strip()]
    if not emails:
        raise SystemExit("no addresses found in " + args.input)

    cache = Cache(args.db)
    verdicts, report = _asyncio.run(run(
        emails, cache, ttl_days=args.ttl,
        use_microsoft=not args.no_microsoft,
        use_smtp=args.smtp,
        nameservers=args.nameserver or None))

    os.makedirs(args.outdir, exist_ok=True)
    fields = ["email", "canonical", "domain", "status", "tier", "route",
              "reason", "suggestion"]

    def write(path, subset):
        with open(path, "w", encoding="utf-8", newline="") as handle:
            writer = csv.writer(handle)
            writer.writerow(fields)
            for v in subset:
                writer.writerow([v.email, v.canonical, v.domain, v.status,
                                 v.tier, v.route, v.reason, v.suggestion])

    billable = [v for v in verdicts if v.disposition == TO_VENDOR]
    siphoned = [v for v in verdicts if v.disposition == SIPHONED]
    review = [v for v in verdicts if v.disposition == REVIEW]

    vendor_path = os.path.join(args.outdir, "send_to_clearout.csv")
    write(vendor_path, billable)
    write(os.path.join(args.outdir, "resolved_by_engine.csv"), siphoned)
    write(os.path.join(args.outdir, "review.csv"), review)

    # Tier 3 needs a clean sending IP, so unless it ran here, export the
    # subset it could answer for and let the server do that pass.
    smtp_path = None
    if not args.smtp:
        from .providers import GOOGLE_CONSUMER as _GC
        from .providers import SELF_HOSTED as _SH
        needs_smtp = [v for v in billable if v.route in (_SH, _GC)]
        if needs_smtp:
            smtp_path = os.path.join(args.outdir, "needs_smtp.csv")
            write(smtp_path, needs_smtp)

    print()
    print("=" * 58)
    print("  input rows            : %d" % report.total)
    print("  unique addresses      : %d" % report.unique)
    print("  resolved by engine    : %d" % report.siphoned)
    print("  needs human review    : %d" % len(review))
    print("  SEND TO CLEAROUT      : %d" % report.billable)
    if report.unique:
        saved = report.unique - report.billable
        print("  credits avoided       : %d of %d unique (%.0f%%)  ~%s%.2f"
              % (saved, report.unique, 100.0 * saved / report.unique,
                 args.currency, saved * args.credit_price))
    print("=" * 58)

    if report.by_tier:
        print("\n  resolved by tier:")
        for tier, n in report.by_tier.most_common():
            print("    %5d  %s" % (n, tier))
    if report.by_route:
        print("\n  how the remaining domains are hosted:")
        for route, n in report.by_route.most_common():
            print("    %5d  %s" % (n, route))

    # If the input carried Clearout verdicts, grade ourselves against them.
    truth_index = _detect_column(_header, ["status", "clearout_status",
                                           "result", "verdict"])
    if truth_index is not None:
        truth = {}
        for row in body:
            if len(row) > max(index, truth_index):
                p = parse(row[index])
                if p.canonical:
                    truth[p.canonical] = (row[truth_index] or "").strip().lower()
        matrix = Counter()
        for v in siphoned:
            actual = truth.get(v.canonical)
            if actual and v.status in ("valid", "invalid"):
                matrix[(v.status, actual)] += 1
        if matrix:
            print("\n  === graded against the vendor verdicts in your file ===")
            wrong = 0
            for (ours, theirs), n in sorted(matrix.items(), key=lambda kv: -kv[1]):
                flag = ""
                if ours == "valid" and theirs == "invalid":
                    flag = "   <-- WOULD BOUNCE"
                    wrong += n
                elif ours == "invalid" and theirs == "valid":
                    flag = "   <-- lost a real lead"
                    wrong += n
                print("    %-9s vs %-9s %d%s" % (ours, theirs, n, flag))
            total = sum(matrix.values())
            print("\n    precision on siphoned addresses: %.1f%% (%d wrong of %d)"
                  % (100.0 * (total - wrong) / total, wrong, total))

    print("\n  wrote %s" % vendor_path)
    print("        %s" % os.path.join(args.outdir, "resolved_by_engine.csv"))
    if smtp_path:
        print("        %s   <-- run this through server/smtp_tier.py on the "
              "server, then: prefilter merge-smtp" % smtp_path)
    cache.close()
    return 0


def cmd_merge_smtp(args: argparse.Namespace) -> int:
    """Fold server-side SMTP verdicts back in and rebuild the vendor list."""
    cache = Cache(args.db)
    _h, body, index = _read_column(args.results, ["email"], None)
    header = _h
    status_index = _detect_column(header, ["status"])
    detail_index = _detect_column(header, ["detail"])
    if status_index is None:
        raise SystemExit("no status column in " + args.results)

    proven = Counter()
    banked = 0
    for row in body:
        if len(row) <= max(index, status_index):
            continue
        parsed = parse(row[index])
        if not parsed.canonical:
            continue
        status = (row[status_index] or "").strip().lower()
        proven[status] += 1
        # Only proven verdicts are banked. policy/catch_all/unknown are not
        # findings, they are the absence of one.
        if status in ("valid", "invalid"):
            detail = (row[detail_index] if detail_index is not None
                      and len(row) > detail_index else "")
            cache.put(parsed.canonical, row[index], "smtp", status, detail)
            banked += 1
    cache.commit()

    print("SMTP tier results:")
    for status, n in proven.most_common():
        print("  %5d  %s" % (n, status))
    print("\n  banked %d proven verdicts into the cache" % banked)
    print("  re-run `prefilter engine` on your list -- these now resolve from "
          "cache and drop out of the Clearout upload")
    cache.close()
    return 0


def cmd_probe(args: argparse.Namespace) -> int:
    """Probe mailboxes over SMTP and grade the result against cached verdicts."""
    import asyncio as _asyncio

    from .smtp_check import BLOCKED, ProbeConfig, probe_all

    cache = Cache(args.db)
    truth: Dict[str, str] = {
        row["canonical"]: row["status"]
        for row in cache.conn.execute(
            "SELECT canonical, status FROM verdicts WHERE source = 'clearout'")
    }
    if args.input:
        _, body, index = _read_column(
            args.input, ["email", "email address", "e-mail", "mail"], args.column)
        targets = [parse(r[index]).canonical for r in body
                   if len(r) > index and (r[index] or "").strip()]
        targets = [t for t in targets if t]
    else:
        targets = sorted(truth)
    if args.limit:
        targets = targets[:args.limit]
    if not targets:
        raise SystemExit("nothing to probe -- ingest some verdicts first")

    config = ProbeConfig(helo=args.helo, mail_from=args.mail_from,
                         timeout=args.timeout,
                         per_domain_delay=args.delay,
                         domain_concurrency=args.concurrency)

    print("probing {0} addresses across {1} domains".format(
        len(targets), len({t.rsplit('@', 1)[1] for t in targets})))
    print("HELO {0} / MAIL FROM <{1}>".format(config.helo, config.mail_from))
    print()

    def progress(done: int, total: int, domain: str) -> None:
        if done % 10 == 0 or done == total:
            print("  {0}/{1} domains".format(done, total), file=sys.stderr)

    results = _asyncio.run(probe_all(targets, config, progress))

    stages: Counter = Counter()
    statuses: Counter = Counter()
    matrix: Counter = Counter()
    for result in results:
        stages[result.stage or "?"] += 1
        statuses[result.status] += 1
        vendor = truth.get(result.email)
        if vendor:
            matrix[(result.status, vendor)] += 1

    print()
    print("=== where the conversation ended ===")
    for stage, count in stages.most_common():
        print("  {0:>6}  {1}".format(count, stage))
    print()
    print("=== our verdicts ===")
    for status, count in statuses.most_common():
        print("  {0:>6}  {1}".format(count, status))

    blocked = statuses.get(BLOCKED, 0)
    if results:
        print()
        print("  refused before we could ask about the mailbox: {0}/{1} ({2:.0f}%)"
              .format(blocked, len(results), 100.0 * blocked / len(results)))

    if matrix:
        print()
        print("=== ours vs Clearout ===")
        print("  {0:<12} {1:<12} {2}".format("ours", "clearout", "n"))
        for (ours, vendor), count in sorted(matrix.items(), key=lambda kv: -kv[1]):
            flag = ""
            if ours == VALID_LABEL and vendor == "invalid":
                flag = "   <-- WOULD BOUNCE"
            elif ours == "invalid" and vendor == "valid":
                flag = "   <-- lost a real lead"
            print("  {0:<12} {1:<12} {2}{3}".format(ours, vendor, count, flag))

        decisive = sum(n for (o, v), n in matrix.items()
                       if o in ("valid", "invalid") and v in ("valid", "invalid"))
        agreed = sum(n for (o, v), n in matrix.items()
                     if o == v and o in ("valid", "invalid"))
        print()
        if decisive:
            print("  decisive comparisons : {0}".format(decisive))
            print("  agreement            : {0:.1f}%".format(100.0 * agreed / decisive))
        else:
            print("  no decisive comparisons -- we could not resolve a single "
                  "mailbox either way")

    cache.close()
    return 0


VALID_LABEL = "valid"


def cmd_stats(args: argparse.Namespace) -> int:
    cache = Cache(args.db)
    verdicts, domains = cache.counts()
    print("cached verdicts : {0}".format(verdicts))
    print("cached domains  : {0}".format(domains))
    print()
    breakdown = cache.status_breakdown()
    if breakdown:
        print("by status:")
        for status, count in sorted(breakdown.items(), key=lambda kv: -kv[1]):
            print("  {0:>8}  {1}".format(count, status))

    shadow = cache.shadow_summary()
    print()
    if shadow["total"] == 0:
        print("no shadow comparisons yet -- run with --shadow, then ingest the results")
    else:
        agreement = 100.0 * shadow["agreed"] / shadow["total"]
        print("shadow comparisons : {0}".format(shadow["total"]))
        print("agreement          : {0:.2f}%".format(agreement))
        print("we said valid, Clearout said invalid : {0}   <-- costly, must be 0"
              .format(shadow["false_positive"]))
        print("we said invalid, Clearout said valid : {0}   <-- wasted leads"
              .format(shadow["false_negative"]))
    cache.close()
    return 0


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(prog="prefilter",
                                     description="Shrink a list before paying to verify it.")
    parser.add_argument("--db", default=DEFAULT_DB, help="SQLite cache path")
    sub = parser.add_subparsers(dest="command", required=True)

    run = sub.add_parser("run", help="pre-filter a raw list")
    run.add_argument("input")
    run.add_argument("--outdir", default="out")
    run.add_argument("--column", help="email column name or index")
    run.add_argument("--ttl", type=int, default=90, help="verdict TTL in days")
    run.add_argument("--domain-ttl", type=int, default=30)
    run.add_argument("--concurrency", type=int, default=25)
    run.add_argument("--timeout", type=float, default=10.0)
    run.add_argument("--drop-role", action="store_true",
                     help="drop info@/sales@ style addresses (off by default: "
                          "role accounts are often the B2B target)")
    run.add_argument("--disposable-file", help="newline-separated domain list")
    run.add_argument("--nameserver", action="append",
                     help="override resolver (repeatable)")
    run.add_argument("--shadow", action="store_true",
                     help="upload the full list so Clearout grades our drops")
    run.add_argument("--credit-price", type=float, default=0.004)
    run.add_argument("--currency", default="$")
    run.set_defaults(func=cmd_run)

    ingest = sub.add_parser("ingest", help="load verified result CSVs into the cache")
    ingest.add_argument("input", nargs="+", help="file paths or globs")
    ingest.add_argument("--column", help="email column name or index")
    ingest.add_argument("--status-column", help="status column name or index")
    ingest.add_argument("--status-map", action="append", metavar="LABEL=STATUS",
                        help="map an export's own label onto ours, e.g. "
                             "--status-map \"Not Validated=invalid\" (repeatable)")
    ingest.add_argument("--empty-status", metavar="STATUS",
                        help="what a blank status cell means. Left unset a blank "
                             "stays 'unknown' and gets re-verified, which is the "
                             "safe default -- only set this if the export "
                             "guarantees every row was processed")
    ingest.add_argument("--dry-run", action="store_true",
                        help="report what would be imported without writing")
    ingest.set_defaults(func=cmd_ingest)

    engine = sub.add_parser(
        "engine", help="run the full tiered engine (recommended)")
    engine.add_argument("input")
    engine.add_argument("--outdir", default="out")
    engine.add_argument("--column", help="email column name or index")
    engine.add_argument("--ttl", type=int, default=90)
    engine.add_argument("--no-microsoft", action="store_true",
                        help="skip the Microsoft HTTPS tier")
    engine.add_argument("--smtp", action="store_true",
                        help="enable SMTP probing. Only meaningful from a host "
                             "with a clean, forward-confirmed sending IP -- from "
                             "a home connection it produces false rejections")
    engine.add_argument("--nameserver", action="append")
    engine.add_argument("--credit-price", type=float, default=0.004)
    engine.add_argument("--currency", default="$")
    engine.set_defaults(func=cmd_engine)

    merge = sub.add_parser("merge-smtp",
                           help="fold server-side SMTP results back into the cache")
    merge.add_argument("results", help="smtp_results.csv produced on the server")
    merge.set_defaults(func=cmd_merge_smtp)

    probe = sub.add_parser("probe", help="SMTP-probe mailboxes and grade against the cache")
    probe.add_argument("input", nargs="?", help="file to probe (default: cached Clearout verdicts)")
    probe.add_argument("--column", help="email column name or index")
    probe.add_argument("--limit", type=int, help="probe only the first N")
    probe.add_argument("--helo", default="tradegeniusglobal.com")
    probe.add_argument("--mail-from", default="postmaster@tradegeniusglobal.com")
    probe.add_argument("--timeout", type=float, default=12.0)
    probe.add_argument("--delay", type=float, default=1.0,
                       help="seconds between probes to the same server")
    probe.add_argument("--concurrency", type=int, default=12,
                       help="distinct domains probed at once")
    probe.set_defaults(func=cmd_probe)

    stats = sub.add_parser("stats", help="cache size and shadow accuracy")
    stats.set_defaults(func=cmd_stats)
    return parser


def main(argv: Optional[List[str]] = None) -> int:
    args = build_parser().parse_args(argv)
    return args.func(args)


if __name__ == "__main__":
    raise SystemExit(main())
